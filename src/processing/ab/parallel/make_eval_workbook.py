#!/usr/bin/env python3
"""Build a blind A/B human-evaluation workbook for the BBC alignment experiment.

The workbook UI is in RUSSIAN (the evaluators' language): headers, dropdown values,
instructions and the summary sheet. Internal scoring (score_eval_workbook.py) maps
the Russian rating values back to canonical buckets.

Reads the per-model alignment JSON produced by align_bbc.py, reconstructs each
pair's text from the source word tokens, and writes an .xlsx with one sheet per
model (anonymized as "Модель A" / "Модель B"), dropdown rating cells, and a summary
sheet that auto-computes alignment accuracy. The A/B -> real-model mapping is written
to a SEPARATE key file so it is not shared with evaluators.

What evaluators judge: ALIGNMENT correctness (does the Abkhaz span correspond to
the Russian span; is the split sensible) — not translation quality, which is a
property of the shared source text and identical across models.

Usage:
    python make_eval_workbook.py --episode 5
"""
import argparse
import datetime as dt
import json
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import align_bbc as a  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

MODELS = {
    5: ["anthropic/claude-opus-4.7", "google/gemini-3.1-pro-preview"],
}

# Anonymized sheet labels + Russian rating vocabulary (shared with the scorer).
LABELS = ["Модель A", "Модель B"]
RATING_CORRECT = "Верно"
RATING_PARTIAL = "Частично"
RATING_WRONG = "Неверно"
ALIGN_CHOICES = [RATING_CORRECT, RATING_PARTIAL, RATING_WRONG]
ISSUE_CHOICES = [
    "Неправильное сопоставление",
    "Плохая сегментация",
    "Ошибка перевода/источника",
    "Другое",
]

HEADERS = ["№", "Абхазский (ab)", "Русский (ru)", "Соответствие", "Тип ошибки", "Комментарии"]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GAP_FILL = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def load_pairs(model: str, ep: int, ab_words, ru_words):
    """Return list of (idx, ab_text, ru_text, is_gap) for every link."""
    tag = a.slug(model)
    data = json.loads((a.OUT_DIR / f"ep{ep}_{tag}_alignment.json").read_text(encoding="utf-8"))
    rows = []
    for i, link in enumerate(data["links"], 1):
        ab = a.span_to_text(link.get("ab"), ab_words)
        ru = a.span_to_text(link.get("ru"), ru_words)
        rows.append((i, ab, ru, not (ab and ru)))
    return rows


def build_model_sheet(ws, rows):
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, ab, ru, is_gap in rows:
        ws.append([idx, ab or "— (нет абхазского)", ru or "— (нет русского)", "", "", ""])
        r = ws.max_row
        ws.cell(r, 2).alignment = WRAP
        ws.cell(r, 3).alignment = WRAP
        ws.cell(r, 1).alignment = TOP
        if is_gap:  # highlight intentional one-sided rows (e.g. credits)
            for c in range(1, 7):
                ws.cell(r, c).fill = GAP_FILL

    last = ws.max_row
    dv_align = DataValidation(type="list", formula1='"%s"' % ",".join(ALIGN_CHOICES), allow_blank=True)
    dv_issue = DataValidation(type="list", formula1='"%s"' % ",".join(ISSUE_CHOICES), allow_blank=True)
    ws.add_data_validation(dv_align)
    ws.add_data_validation(dv_issue)
    dv_align.add(f"D2:D{last}")
    dv_issue.add(f"E2:E{last}")

    widths = {"A": 5, "B": 60, "C": 60, "D": 14, "E": 26, "F": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def build_summary(ws, labels):
    ws.append(["Точность выравнивания — вычисляется автоматически по столбцам оценок"])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])
    ws.append([""] + labels)
    for c in range(2, 2 + len(labels)):
        ws.cell(3, c).font = Font(bold=True)
    cf = '\'{s}\'!D2:D100000'  # rating column on each model sheet
    metrics = [
        ("Всего пар", '=COUNTA(\'{s}\'!A2:A100000)'),
        ("Оценено", f'=COUNTIF({cf},"<>")'),
        (RATING_CORRECT, f'=COUNTIF({cf},"{RATING_CORRECT}")'),
        (RATING_PARTIAL, f'=COUNTIF({cf},"{RATING_PARTIAL}")'),
        (RATING_WRONG, f'=COUNTIF({cf},"{RATING_WRONG}")'),
        (f"Точность ({RATING_CORRECT} / Оценено)",
         f'=IFERROR(COUNTIF({cf},"{RATING_CORRECT}")/COUNTIF({cf},"<>"),"")'),
        (f"Мягкая оценка (({RATING_CORRECT}+0.5×{RATING_PARTIAL})/Оценено)",
         f'=IFERROR((COUNTIF({cf},"{RATING_CORRECT}")+0.5*COUNTIF({cf},"{RATING_PARTIAL}"))'
         f'/COUNTIF({cf},"<>"),"")'),
    ]
    for label, fmla in metrics:
        ws.append([label] + [fmla.format(s=s) for s in labels])
        r = ws.max_row
        is_acc = label.startswith(("Точность", "Мягкая"))
        ws.cell(r, 1).font = Font(bold=is_acc)
        if is_acc:
            for c in range(2, 2 + len(labels)):
                ws.cell(r, c).number_format = "0.0%"
    ws.column_dimensions["A"].width = 46
    for c in range(2, 2 + len(labels)):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 14


def build_instructions(ws, ep):
    lines = [
        (f"BBC серия {ep} — выравнивание абхазского↔русского: экспертная оценка", True, 14),
        ("", False, 11),
        ("Вы оцениваете КАЧЕСТВО ВЫРАВНИВАНИЯ, а не качество перевода.", True, 11),
        ("Абхазский и русский текст взяты из одного источника; оба варианта используют один и тот же "
         "текст. Различается только то, КАК текст разбит на части и сопоставлен. Оцените, "
         "соответствуют ли друг другу абхазская и русская стороны в каждой паре.", False, 11),
        ("", False, 11),
        ("Листы «Модель A» и «Модель B» — две анонимные системы. Оценивайте каждую независимо.", False, 11),
        ("", False, 11),
        ("Столбец «Соответствие» (выпадающий список):", True, 11),
        (f"  {RATING_CORRECT}    — стороны сопоставлены правильно и разумно разбиты. Естественное "
         "сжатие перевода (субтитры короче исходного текста) — это всё равно «Верно».", False, 11),
        (f"  {RATING_PARTIAL} — партнёр верный, но границы неудачны: фраза попала не в ту пару, "
         "лишняя или потеряна ИЗ-ЗА разреза (а не из-за самого перевода).", False, 11),
        (f"  {RATING_WRONG}  — стороны не соответствуют друг другу (не тот партнёр).", False, 11),
        ("", False, 11),
        ("Важно — про «потерянный смысл». Если фраза есть на одной стороне, но отсутствует на другой, "
         "определите причину:", True, 11),
        ("  • её «отрезало» выравнивание — фраза есть рядом (в соседней паре) или должна была войти "
         "сюда → «Частично» + «Плохая сегментация» (это ошибка модели).", False, 11),
        ("  • её просто НЕТ в самом переводе — русские субтитры короче абхазского текста, сопоставлять "
         "нечего → это НЕ ошибка выравнивания: ставьте «Верно» и при необходимости отметьте "
         "«Ошибка перевода/источника». Модель не может сопоставить то, чего нет.", False, 11),
        ("", False, 11),
        ("Жёлтые строки — это намеренные ПРОПУСКИ (одна сторона пустая, например титры на экране "
         "без перевода). Поставьте «Верно», если строку правильно оставить без пары, иначе «Неверно».", False, 11),
        ("", False, 11),
        ("Столбец «Тип ошибки» (необязательно): если оценка не «Верно», выберите основную проблему:", True, 11),
        ("  • Неправильное сопоставление — абхазская и русская стороны связаны неверно, они о РАЗНОМ "
         "(не тот партнёр). Границы могут быть верными, но пара сопоставляет не те фрагменты. Обычно "
         "вместе с оценкой «Неверно». Это самая серьёзная ошибка — такая пара портит корпус.", False, 11),
        ("  • Плохая сегментация — стороны соответствуют нужному месту, но текст РАЗРЕЗАН неудачно: "
         "предложение разбито не там, два разных предложения объединены в одну пару, или лишняя фраза "
         "попала из соседней пары (правильный партнёр, но неверные «ножницы»). Обычно вместе с «Частично».", False, 11),
        ("  • Ошибка перевода/источника — проблема в самом исходном тексте (опечатка, пропуск, ошибка "
         "перевода в оригинале), а НЕ в выравнивании.", False, 11),
        ("  • Другое — иная проблема; поясните в столбце «Комментарии».", False, 11),
        ("", False, 11),
        ("Подсказка: «не тот партнёр» → Неправильное сопоставление;  «правильный партнёр, но плохо "
         "разрезано» → Плохая сегментация.", False, 11),
        ("", False, 11),
        ("Столбец «Комментарии»: любые замечания, полезные для итогового решения.", False, 11),
        ("", False, 11),
        ("Лист «Итоги» автоматически вычисляет точность каждой модели по мере заполнения оценок. "
         "Целевая точность корпуса — 95%.", False, 11),
    ]
    for text, bold, size in lines:
        ws.append([text])
        ws.cell(ws.max_row, 1).font = Font(bold=bold, size=size)
    ws.column_dimensions["A"].width = 115


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--episode", type=int, default=5)
    ap.add_argument("--seed", type=int, default=None, help="fix the A/B shuffle for reproducibility")
    args = ap.parse_args()
    ep = args.episode

    ab_name, ru_name = a.EPISODES[ep]
    ab_words = a.tokenize(a.RAW_BBC / ab_name)
    ru_words = a.tokenize(a.RAW_BBC / ru_name)

    models = list(MODELS[ep])
    random.Random(args.seed).shuffle(models)  # models[0] -> Модель A, models[1] -> Модель B
    mapping = dict(zip(LABELS, models))

    wb = Workbook()
    build_instructions(wb.active, ep)
    wb.active.title = "Инструкция"
    for label, model in mapping.items():
        ws = wb.create_sheet(label)
        build_model_sheet(ws, load_pairs(model, ep, ab_words, ru_words))
    build_summary(wb.create_sheet("Итоги"), LABELS)

    out_xlsx = a.OUT_DIR / f"ep{ep}_human_eval.xlsx"
    wb.save(out_xlsx)

    key = a.OUT_DIR / f"ep{ep}_human_eval_KEY.txt"
    key.write_text(
        "BLIND KEY — keep private, do NOT share with evaluators.\n"
        f"Generated {dt.date.today().isoformat()}  (seed={args.seed})\n\n"
        + "\n".join(f"{label} = {model}" for label, model in mapping.items())
        + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {out_xlsx}")
    print(f"Wrote {key}  (mapping kept out of the workbook)")
    for label, model in mapping.items():
        print(f"  {label} -> {model}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
