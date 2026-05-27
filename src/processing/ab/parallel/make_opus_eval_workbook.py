#!/usr/bin/env python3
"""Build an Opus-only human-evaluation workbook spanning ALL aligned BBC episodes.

Unlike make_eval_workbook.py (a blind A/B Opus-vs-Gemini test on a single episode),
this is a single-model (Opus subagent) review across every aligned episode. It samples
a CONTIGUOUS window of pairs from each episode — adjacency is deliberate: consecutive
pairs are what let an evaluator catch bad SEGMENTATION (a sentence split across two
pairs, or text that leaked in from a neighbour). Gap rows that fall inside the window
are kept (highlighted) so the run stays truly adjacent.

Layout: one sheet per episode ("ВВС N"), N consecutive pairs each, plus a Russian
instructions sheet and an auto-computing summary (per-episode + overall accuracy).
The UI/vocabulary is shared with make_eval_workbook.py, so score_eval_workbook-style
scoring (COUNTIF on column D) works per sheet.

Usage (system python lacks openpyxl; use the uv cache on PYTHONPATH):
    PYTHONPATH="$(echo ~/.cache/uv/archive-v0/*/ | tr ' ' ':')" \
        python3 make_opus_eval_workbook.py --pairs 30 --seed 0
"""
import argparse
import datetime as dt
import json
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import align_bbc as a  # noqa: E402
from make_eval_workbook import (  # noqa: E402  (reuse shared UI + sheet builders)
    HEADERS, RATING_CORRECT, RATING_PARTIAL, RATING_WRONG,
    build_model_sheet,
)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

# Opus subagent is the production aligner for every episode (see align_bbc / commit history).
MODEL_TAG = "anthropic-claude-opus-4-7-subagent"
EPISODES = [1, 2, 4, 5, 7, 9, 10]  # every episode with an Abkhaz source, now aligned


def load_pairs(ep: int, ab_words, ru_words):
    """Return [(link_idx, ab_text, ru_text, is_gap)] for every link in the episode."""
    data = json.loads((a.OUT_DIR / f"ep{ep}_{MODEL_TAG}_alignment.json").read_text(encoding="utf-8"))
    rows = []
    for i, link in enumerate(data["links"], 1):
        ab = a.span_to_text(link.get("ab"), ab_words)
        ru = a.span_to_text(link.get("ru"), ru_words)
        rows.append((i, ab, ru, not (ab and ru)))
    return rows


def adjacent_window(rows, n_pairs: int, rng: random.Random):
    """Smallest CONTIGUOUS slice of rows holding n_pairs paired rows (gaps kept inside).

    Adjacency is the point: it exposes segmentation errors that span neighbouring pairs.
    The slice keeps each link's original № so a flagged pair traces back to the full
    alignment. Start is chosen randomly among valid positions (reproducible via seed).
    """
    pair_pos = [i for i, r in enumerate(rows) if not r[3]]
    if len(pair_pos) <= n_pairs:
        return rows
    sp = rng.randint(0, len(pair_pos) - n_pairs)
    return rows[pair_pos[sp]: pair_pos[sp + n_pairs - 1] + 1]


def build_instructions(ws, n_pairs: int):
    lines = [
        ("BBC: выравнивание абхазского↔русского (модель Opus) — экспертная оценка", True, 14),
        ("", False, 11),
        ("Вы оцениваете КАЧЕСТВО ВЫРАВНИВАНИЯ, а не качество перевода.", True, 11),
        ("Абхазский и русский текст взяты из одного источника. Различается только то, КАК текст "
         "разбит на части и сопоставлен. Оцените, соответствуют ли друг другу абхазская и русская "
         "стороны в каждой паре.", False, 11),
        ("", False, 11),
        (f"Каждый лист — отдельная серия BBC; на нём {n_pairs} ИДУЩИХ ПОДРЯД пар из этой серии. "
         "Пары намеренно соседние: так видно ошибки СЕГМЕНТАЦИИ — например, одно предложение, "
         "разрезанное на две пары, или фрагмент, попавший из соседней пары.", False, 11),
        ("", False, 11),
        ("Столбец «Соответствие» (выпадающий список):", True, 11),
        (f"  {RATING_CORRECT}    — стороны сопоставлены правильно и разумно разбиты. Естественное "
         "сжатие перевода (субтитры короче исходного текста) — это всё равно «Верно».", False, 11),
        (f"  {RATING_PARTIAL} — партнёр верный, но границы неудачны: фраза попала не в ту пару, "
         "лишняя или потеряна ИЗ-ЗА разреза (а не из-за самого перевода).", False, 11),
        (f"  {RATING_WRONG}  — стороны не соответствуют друг другу (не тот партнёр).", False, 11),
        ("", False, 11),
        ("Важно — про «потерянный смысл». Если фраза есть на одной стороне, но отсутствует на "
         "другой, определите причину:", True, 11),
        ("  • её «отрезало» выравнивание — фраза есть рядом (в соседней паре) или должна была войти "
         "сюда → «Частично» + «Плохая сегментация» (это ошибка модели).", False, 11),
        ("  • её просто НЕТ в самом переводе — русские субтитры короче абхазского текста → это НЕ "
         "ошибка выравнивания: ставьте «Верно» и при необходимости «Ошибка перевода/источника».", False, 11),
        ("", False, 11),
        ("Жёлтые строки — намеренные ПРОПУСКИ (одна сторона пустая, например титры на экране без "
         "перевода). Поставьте «Верно», если строку правильно оставить без пары, иначе «Неверно».", False, 11),
        ("", False, 11),
        ("Столбец «Тип ошибки» (необязательно, если оценка не «Верно»):", True, 11),
        ("  • Неправильное сопоставление — стороны о РАЗНОМ (не тот партнёр). Самая серьёзная ошибка.", False, 11),
        ("  • Плохая сегментация — правильный партнёр, но текст разрезан неудачно (не там граница, "
         "два предложения слиты, лишняя фраза из соседней пары).", False, 11),
        ("  • Ошибка перевода/источника — проблема в самом исходном тексте, а НЕ в выравнивании.", False, 11),
        ("  • Другое — поясните в «Комментарии».", False, 11),
        ("", False, 11),
        ("Лист «Итоги» автоматически считает точность по каждой серии и в целом. Цель корпуса — 95%.", False, 11),
        ("", False, 11),
        ("Примечание: № в первом столбце — это номер пары в ПОЛНОМ выравнивании серии (не подряд "
         "1..N), чтобы потом найти проблемную пару в исходных данных.", False, 10),
    ]
    for text, bold, size in lines:
        ws.append([text])
        ws.cell(ws.max_row, 1).font = Font(bold=bold, size=size)
    ws.column_dimensions["A"].width = 115


def build_summary(ws, sheet_labels):
    """Per-episode + overall accuracy, computed live from each episode sheet's column D."""
    ws.append(["Точность выравнивания (Opus) — вычисляется автоматически"])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Серия", "Всего", "Оценено", RATING_CORRECT, RATING_PARTIAL, RATING_WRONG,
               "Точность", "Мягкая оценка"])
    for c in range(1, 9):
        ws.cell(3, c).font = Font(bold=True)

    def d(s):  # rating column on a given sheet
        return f"'{s}'!D2:D100000"

    for s in sheet_labels:
        ws.append([
            s,
            f"=COUNTA('{s}'!A2:A100000)",
            f'=COUNTIF({d(s)},"<>")',
            f'=COUNTIF({d(s)},"{RATING_CORRECT}")',
            f'=COUNTIF({d(s)},"{RATING_PARTIAL}")',
            f'=COUNTIF({d(s)},"{RATING_WRONG}")',
            f'=IFERROR(COUNTIF({d(s)},"{RATING_CORRECT}")/COUNTIF({d(s)},"<>"),"")',
            f'=IFERROR((COUNTIF({d(s)},"{RATING_CORRECT}")+0.5*COUNTIF({d(s)},"{RATING_PARTIAL}"))'
            f'/COUNTIF({d(s)},"<>"),"")',
        ])
        r = ws.max_row
        ws.cell(r, 7).number_format = "0.0%"
        ws.cell(r, 8).number_format = "0.0%"

    # overall row sums the per-episode count columns
    first, last = 4, ws.max_row
    tot = ws.max_row + 1
    ws.append([
        "ВСЕГО",
        f"=SUM(B{first}:B{last})", f"=SUM(C{first}:C{last})", f"=SUM(D{first}:D{last})",
        f"=SUM(E{first}:E{last})", f"=SUM(F{first}:F{last})",
        f'=IFERROR(D{tot}/C{tot},"")', f'=IFERROR((D{tot}+0.5*E{tot})/C{tot},"")',
    ])
    for c in range(1, 9):
        ws.cell(tot, c).font = Font(bold=True)
    ws.cell(tot, 7).number_format = "0.0%"
    ws.cell(tot, 8).number_format = "0.0%"
    ws.column_dimensions["A"].width = 12
    for col in "BCDEF":
        ws.column_dimensions[col].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--pairs", type=int, default=30, help="paired rows per episode")
    ap.add_argument("--seed", type=int, default=0, help="fix the window starts for reproducibility")
    args = ap.parse_args()
    rng = random.Random(args.seed)

    wb = Workbook()
    build_instructions(wb.active, args.pairs)
    wb.active.title = "Инструкция"

    sheet_labels = []
    print(f"Sampling {args.pairs} adjacent pairs/episode (seed={args.seed}):")
    for ep in EPISODES:
        ab_name, ru_name = a.EPISODES[ep]
        ab_words = a.tokenize(a.RAW_BBC / ab_name)
        ru_words = a.tokenize(a.RAW_BBC / ru_name)
        rows = load_pairs(ep, ab_words, ru_words)
        window = adjacent_window(rows, args.pairs, rng)
        label = f"ВВС {ep}"
        sheet_labels.append(label)
        build_model_sheet(wb.create_sheet(label), window)
        n_pair = sum(1 for r in window if not r[3])
        n_gap = len(window) - n_pair
        lo, hi = window[0][0], window[-1][0]
        print(f"  {label}: links №{lo}–{hi}  ({n_pair} pairs + {n_gap} gap rows)")

    build_summary(wb.create_sheet("Итоги"), sheet_labels)

    out = a.OUT_DIR / "bbc_opus_human_eval.xlsx"
    wb.save(out)
    print(f"\nWrote {out}  ({len(sheet_labels)} episode sheets, "
          f"{len(sheet_labels) * args.pairs} pairs total)")
    print(f"Generated {dt.date.today().isoformat()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
