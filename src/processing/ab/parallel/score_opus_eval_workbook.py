#!/usr/bin/env python3
"""Score the filled-in Opus all-episode workbook (bbc_opus_human_eval.xlsx).

Per-episode and overall alignment accuracy (strict + lenient) with a Wilson 95%
interval, an issue-type tally, and the flagged (Частично/Неверно) pairs for
qualitative review. Reuses the sheet-scoring logic from score_eval_workbook.py.

Usage (system python lacks openpyxl; use the uv cache on PYTHONPATH):
    PYTHONPATH="$(echo ~/.cache/uv/archive-v0/*/ | tr ' ' ':')" \
        python3 score_opus_eval_workbook.py
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import align_bbc as a  # noqa: E402
from score_eval_workbook import score_sheet, wilson, fmt_pct  # noqa: E402
from make_opus_eval_workbook import EPISODES  # noqa: E402

from openpyxl import load_workbook  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--show-flagged", type=int, default=8,
                    help="flagged (Частично/Неверно) pairs to list per episode")
    args = ap.parse_args()

    xlsx = a.OUT_DIR / "bbc_opus_human_eval.xlsx"
    if not xlsx.exists():
        sys.exit(f"No workbook at {xlsx}")
    wb = load_workbook(xlsx)
    labels = [f"ВВС {ep}" for ep in EPISODES if f"ВВС {ep}" in wb.sheetnames]

    report = ["=== BBC Opus alignment — human-evaluation results ===\n"]
    agg = {"Correct": 0, "Partial": 0, "Wrong": 0}
    issues_all: dict[str, int] = {}
    target = 0.95

    for label in labels:
        res = score_sheet(wb[label])
        c = res["counts"]
        report.append(f"## {label}")
        if res["rated"] == 0:
            report.append("   (no ratings filled in yet)\n")
            continue
        for k in agg:
            agg[k] += c[k]
        for k, v in res["issues"].items():
            issues_all[k] = issues_all.get(k, 0) + v
        report.append(f"   rated {res['rated']}/{res['total_pairs']}  |  "
                      f"Верно {c['Correct']}  Частично {c['Partial']}  Неверно {c['Wrong']}")
        report.append(f"   strict {fmt_pct(res['strict_accuracy'])} "
                      f"(95% CI {fmt_pct(res['ci95'][0])}–{fmt_pct(res['ci95'][1])})  |  "
                      f"lenient {fmt_pct(res['lenient_accuracy'])}")
        report.append("")

    rated = sum(agg.values())
    if rated:
        strict = agg["Correct"] / rated
        lenient = (agg["Correct"] + 0.5 * agg["Partial"]) / rated
        lo, hi = wilson(agg["Correct"], rated)
        report.append("## OVERALL (all episodes)")
        report.append(f"   rated {rated}  |  Верно {agg['Correct']}  "
                      f"Частично {agg['Partial']}  Неверно {agg['Wrong']}")
        report.append(f"   strict {fmt_pct(strict)} (95% CI {fmt_pct(lo)}–{fmt_pct(hi)})  |  "
                      f"lenient {fmt_pct(lenient)}")
        report.append(f"   vs 95% target: {'MEETS' if lo >= target else 'below (CI lower bound < 95%)'}")
        if issues_all:
            report.append("   issues: " + ", ".join(
                f"{k}: {v}" for k, v in sorted(issues_all.items(), key=lambda x: -x[1])))
        report.append("")

    for label in labels:
        res = score_sheet(wb[label])
        if not res["flagged"]:
            continue
        report.append(f"--- {label}: flagged (up to {args.show_flagged} of {len(res['flagged'])}) ---")
        for idx, bucket, issue, ab, ru, notes in res["flagged"][: args.show_flagged]:
            report.append(f"  №{idx} {bucket}" + (f" / {issue}" if issue else "")
                          + (f"  — {notes}" if notes else ""))
            report.append(f"      AB: {ab[:90]}")
            report.append(f"      RU: {ru[:90]}")
        report.append("")

    text = "\n".join(report)
    print(text)
    out = a.OUT_DIR / "bbc_opus_eval_results.txt"
    out.write_text(text, encoding="utf-8")
    print(f"(written to {out})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
