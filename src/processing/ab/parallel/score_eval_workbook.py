#!/usr/bin/env python3
"""Score a filled-in blind human-evaluation workbook and reveal the A/B identities.

Reads ep{N}_human_eval.xlsx (rated by evaluators) plus the private
ep{N}_human_eval_KEY.txt, then per model computes alignment accuracy (strict and
lenient) with a Wilson 95% confidence interval, tallies issue types, lists the
flagged (Partial/Wrong) pairs, and declares which real model scored higher.

Usage:
    python score_eval_workbook.py --episode 5
"""
import argparse
import math
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import align_bbc as a  # noqa: E402  (for OUT_DIR)
from make_eval_workbook import (  # noqa: E402  (shared Russian UI vocabulary)
    LABELS, RATING_CORRECT, RATING_PARTIAL, RATING_WRONG,
)

from openpyxl import load_workbook  # noqa: E402

# Russian rating value -> canonical bucket used internally for scoring.
BUCKET = {
    RATING_CORRECT.lower(): "Correct",
    RATING_PARTIAL.lower(): "Partial",
    RATING_WRONG.lower(): "Wrong",
}
VALID = set(BUCKET)


def wilson(k: int, n: int, z: float = 1.96) -> tuple[float, float]:
    """95% Wilson score interval for a proportion k/n. Returns (low, high)."""
    if n == 0:
        return (0.0, 0.0)
    p = k / n
    denom = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denom
    half = (z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))) / denom
    return (max(0.0, center - half), min(1.0, center + half))


def read_key(ep: int) -> dict:
    key_path = a.OUT_DIR / f"ep{ep}_human_eval_KEY.txt"
    mapping = {}
    if key_path.exists():
        for line in key_path.read_text(encoding="utf-8").splitlines():
            if "=" in line and line.split("=", 1)[0].strip() in LABELS:
                label, model = line.split("=", 1)
                mapping[label.strip()] = model.strip()
    return mapping


def score_sheet(ws) -> dict:
    counts = {"Correct": 0, "Partial": 0, "Wrong": 0}
    issues: dict[str, int] = {}
    flagged = []  # (row#, rating, issue, ab, ru, notes)
    total = 0
    for r in range(2, ws.max_row + 1):
        idx = ws.cell(r, 1).value
        if idx in (None, ""):
            continue
        total += 1
        rating = (str(ws.cell(r, 4).value or "")).strip()
        norm = rating.lower()
        if norm not in VALID:
            continue
        bucket = BUCKET[norm]
        counts[bucket] += 1
        if bucket != "Correct":
            issue = (str(ws.cell(r, 5).value or "")).strip()
            if issue:
                issues[issue] = issues.get(issue, 0) + 1
            flagged.append((idx, bucket, issue,
                            str(ws.cell(r, 2).value or ""), str(ws.cell(r, 3).value or ""),
                            str(ws.cell(r, 6).value or "")))
    rated = counts["Correct"] + counts["Partial"] + counts["Wrong"]
    strict = counts["Correct"] / rated if rated else 0.0
    lenient = (counts["Correct"] + 0.5 * counts["Partial"]) / rated if rated else 0.0
    lo, hi = wilson(counts["Correct"], rated)
    return {
        "total_pairs": total,
        "rated": rated,
        "counts": counts,
        "strict_accuracy": strict,
        "lenient_accuracy": lenient,
        "ci95": (lo, hi),
        "issues": issues,
        "flagged": flagged,
    }


def fmt_pct(x: float) -> str:
    return f"{x * 100:.1f}%"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--episode", type=int, default=5)
    ap.add_argument("--show-flagged", type=int, default=10,
                    help="how many flagged (Partial/Wrong) pairs to list per model")
    args = ap.parse_args()
    ep = args.episode

    xlsx = a.OUT_DIR / f"ep{ep}_human_eval.xlsx"
    if not xlsx.exists():
        sys.exit(f"No workbook at {xlsx}")
    wb = load_workbook(xlsx)
    mapping = read_key(ep)  # {"Model A": real, ...}

    results = {}
    for label in LABELS:
        if label in wb.sheetnames:
            results[label] = score_sheet(wb[label])

    report = [f"=== Episode {ep} human-evaluation results ===\n"]
    target = 0.95
    summary_rows = []
    for label, res in results.items():
        real = mapping.get(label, "(unknown — key file missing)")
        c = res["counts"]
        report.append(f"## {label}  =  {real}")
        if res["rated"] == 0:
            report.append("   (no ratings filled in yet)\n")
            continue
        report.append(
            f"   rated {res['rated']}/{res['total_pairs']}  |  "
            f"Correct {c['Correct']}  Partial {c['Partial']}  Wrong {c['Wrong']}"
        )
        report.append(
            f"   strict accuracy:  {fmt_pct(res['strict_accuracy'])}  "
            f"(95% CI {fmt_pct(res['ci95'][0])}–{fmt_pct(res['ci95'][1])})"
        )
        report.append(f"   lenient accuracy: {fmt_pct(res['lenient_accuracy'])}")
        report.append(
            f"   vs 95% target:    {'MEETS' if res['ci95'][0] >= target else 'below (CI lower bound under 95%)'}"
        )
        if res["issues"]:
            tally = ", ".join(f"{k}: {v}" for k, v in sorted(res["issues"].items(), key=lambda x: -x[1]))
            report.append(f"   issue types:      {tally}")
        report.append("")
        summary_rows.append((real, res["strict_accuracy"], res["rated"]))

    # verdict
    rated_results = [(lbl, r) for lbl, r in results.items() if r["rated"] > 0]
    if len(rated_results) == 2:
        (la, ra), (lb, rb) = rated_results
        if ra["strict_accuracy"] != rb["strict_accuracy"]:
            win_lbl, win = (la, ra) if ra["strict_accuracy"] > rb["strict_accuracy"] else (lb, rb)
            report.append(
                f"VERDICT: {mapping.get(win_lbl, win_lbl)} scored higher "
                f"({fmt_pct(win['strict_accuracy'])} strict)."
            )
        else:
            report.append("VERDICT: both models tied on strict accuracy.")
        report.append("")

    # flagged pairs for qualitative review
    for label, res in results.items():
        if not res["flagged"]:
            continue
        report.append(f"--- {label} ({mapping.get(label, '?')}): flagged pairs "
                      f"(showing up to {args.show_flagged} of {len(res['flagged'])}) ---")
        for idx, bucket, issue, ab, ru, notes in res["flagged"][: args.show_flagged]:
            report.append(f"  [{idx}] {bucket}" + (f" / {issue}" if issue else "")
                          + (f"  — {notes}" if notes else ""))
            report.append(f"      AB: {ab[:90]}")
            report.append(f"      RU: {ru[:90]}")
        report.append("")

    text = "\n".join(report)
    print(text)
    out = a.OUT_DIR / f"ep{ep}_eval_results.txt"
    out.write_text(text, encoding="utf-8")
    print(f"(written to {out})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
